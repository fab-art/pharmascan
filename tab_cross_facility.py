""" tab_cross_facility.py — Cross-facility RAMA number fraud detection. """
import io
import math
import re
import streamlit as st
import pandas as pd


def render(tab, df, s, col_map):
    with tab:
        with tab_xfac:
            import math

            # ── CSS for fraud tab cards ───────────────────────────────────────────────
            st.markdown("""
        <style>
        .fraud-card{background:#111720;border:1px solid #1e2a38;border-radius:12px;
                    padding:18px 22px;margin-bottom:14px}
        .fraud-card-red{background:#1a0505;border-color:#7f1d1d}
        .fraud-card-amber{background:#1a1000;border-color:#78350f}
        .fraud-card-blue{background:#030d1a;border-color:#1e3a5f}
        .fraud-card-green{background:#031a0a;border-color:#14532d}
        .badge{display:inline-block;padding:2px 10px;border-radius:20px;
               font-size:11px;font-weight:700;font-family:monospace}
        .badge-red{background:#7f1d1d;color:#fca5a5}
        .badge-amber{background:#78350f;color:#fde68a}
        .badge-green{background:#14532d;color:#86efac}
        .badge-blue{background:#1e3a5f;color:#93c5fd}
        .badge-purple{background:#3b1f7a;color:#c4b5fd}
        .risk-bar-wrap{background:#1e2a38;border-radius:6px;height:10px;width:100%;margin:4px 0}
        .risk-bar{height:10px;border-radius:6px}
        </style>""", unsafe_allow_html=True)

            st.markdown("""
        <div style='font-family:Syne,sans-serif;font-size:26px;font-weight:800;
             color:#e2e8f0;margin-bottom:4px'>
          🔍 Pharmacy Fraud Detection
        </div>
        <div style='color:#64748b;font-size:13px;margin-bottom:20px;font-family:monospace'>
          Identifies pharmacy claims where medicine was dispensed to patients
          with <b style='color:#ef4444'>no verifiable hospital or clinic visit record</b>
          — a primary indicator of fraudulent RSSB reimbursement claims.
        </div>""", unsafe_allow_html=True)

            # ── How it works banner ────────────────────────────────────────────────────
            st.markdown("""
        <div class='fraud-card' style='border-color:#0ea5e9;background:#020e1a;margin-bottom:22px'>
        <div style='font-size:13px;font-weight:700;color:#38bdf8;margin-bottom:10px'>
          📋 Patient Journey & Fraud Logic
        </div>
        <div style='display:flex;gap:8px;align-items:center;flex-wrap:wrap;font-size:12px;
             font-family:monospace;color:#94a3b8'>
          <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
            🏥 Clinic Visit
          </span>
          <span style='color:#475569'>→</span>
          <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
            👨‍⚕️ Doctor consults
          </span>
          <span style='color:#475569'>→</span>
          <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
            📝 Prescription + Voucher
          </span>
          <span style='color:#475569'>→</span>
          <span style='background:#1e3a5f;color:#93c5fd;padding:4px 10px;border-radius:6px'>
            💰 Patient pays 15%
          </span>
          <span style='color:#475569'>→</span>
          <span style='background:#14532d;color:#86efac;padding:4px 10px;border-radius:6px'>
            💊 Pharmacy dispenses
          </span>
          <span style='color:#475569'>→</span>
          <span style='background:#14532d;color:#86efac;padding:4px 10px;border-radius:6px'>
            📤 Pharmacy claims 85% from RSSB
          </span>
        </div>
        <div style='margin-top:10px;font-size:11px;color:#64748b;font-family:monospace'>
          <b style='color:#ef4444'>⚠️ Fraud signal:</b> Pharmacy claims 85% reimbursement for a patient
          who has <i>no hospital/clinic visit record</i> — meaning there was no legitimate
          consultation, no doctor, and likely no real prescription.
          The more facility files uploaded, the more accurate this detection becomes.
        </div>
        </div>""", unsafe_allow_html=True)

            # ── STEP 1: Upload facility files ─────────────────────────────────────────
            st.markdown('<div class="sec-head">📂 Step 1 — Upload Hospital & Clinic Visit Files</div>',
                        unsafe_allow_html=True)
            st.markdown("""
        <div style='font-size:11px;color:#64748b;font-family:monospace;margin-bottom:10px'>
          Upload one or more hospital/clinic Excel files for the <b>same period</b> as the
          pharmacy report. The app auto-detects the verified sheet and all relevant columns.
          <b style='color:#e2e8f0'>The more facilities you upload, the smaller the unverified group becomes.</b>
        </div>""", unsafe_allow_html=True)

            xf_uploads = st.file_uploader(
                "Upload facility files (Excel / CSV) — multiple allowed",
                type=["xlsx","xls","csv"],
                accept_multiple_files=True,
                key="xf_uploads",
            )

            # ── Parser ────────────────────────────────────────────────────────────────
            def _parse_facility(raw_bytes, filename):
                fname = filename.lower()
                try:
                    if fname.endswith(".csv"):
                        raw = pd.read_csv(io.BytesIO(raw_bytes), encoding="utf-8", on_bad_lines="skip")
                        chosen_name, chosen_df = "main", raw
                        header_row = 0
                    else:
                        xl = pd.ExcelFile(io.BytesIO(raw_bytes))
                        chosen_name = None; chosen_df = None
                        for priority in ["after","verified","clean","before","data","invoice report","invoice"]:
                            for sn in xl.sheet_names:
                                if priority in sn.lower():
                                    chosen_name = sn
                                    chosen_df   = xl.parse(sn, header=None)
                                    break
                            if chosen_name: break
                        if chosen_df is None:
                            sizes = {}
                            for sn in xl.sheet_names:
                                try: sizes[sn] = xl.parse(sn,header=None).shape[0]
                                except: pass
                            chosen_name = max(sizes, key=sizes.get)
                            chosen_df   = xl.parse(chosen_name, header=None)

                        # Find header row
                        header_row = 0
                        for i, row in chosen_df.head(20).iterrows():
                            joined = " ".join(str(v).lower() for v in row if pd.notna(v))
                            if any(k in joined for k in ["affil","rama","beneficiary","patient name","voucher"]):
                                header_row = i; break

                    if header_row > 0:
                        chosen_df = pd.read_excel(io.BytesIO(raw_bytes),
                                                  sheet_name=chosen_name, header=header_row)
                    else:
                        if not isinstance(chosen_df.columns[0], str) or chosen_df.columns[0] == 0:
                            chosen_df.columns = chosen_df.iloc[0]
                            chosen_df = chosen_df.iloc[1:].reset_index(drop=True)

                    chosen_df.columns = [str(c).strip() for c in chosen_df.columns]

                    def _find(patterns):
                        for pat in patterns:
                            for c in chosen_df.columns:
                                if re.search(pat, str(c).lower()): return c
                        return None

                    rama_c  = _find([r"affil", r"rama", r"member.*no"])
                    name_c  = _find([r"benefi.*name", r"patient.*name", r"client.*name"])
                    date_c  = _find([r"^date$", r"visit.*date", r"dispensing.*date", r"service.*date"])
                    vou_c   = _find([r"voucher.*id", r"voucher.*ident", r"paper.*code", r"invoice.*no"])
                    total_c = _find([r"total.*amount", r"total.*cost", r"^total$"])
                    doc_c   = _find([r"practitioner", r"prescrib", r"doctor", r"physician", r"medecin"])

                    if not rama_c:
                        return None, f"No RAMA/Affiliation column found in {chosen_name!r}"

                    # Drop footer/total rows
                    no_col = chosen_df.columns[0]
                    chosen_df = chosen_df[pd.to_numeric(chosen_df[no_col], errors="coerce").notna()].copy()

                    out = pd.DataFrame()
                    out["_rama"]       = chosen_df[rama_c].astype(str).str.strip().str.upper()
                    out["_name"]       = chosen_df[name_c].fillna("").astype(str).str.strip() if name_c else ""
                    out["_date"]       = pd.to_datetime(chosen_df[date_c], errors="coerce") if date_c else pd.NaT
                    out["voucher_id"]  = chosen_df[vou_c].astype(str).str.strip() if vou_c else ""
                    out["total"]       = pd.to_numeric(chosen_df[total_c], errors="coerce").fillna(0) if total_c else 0
                    out["doctor"]      = chosen_df[doc_c].fillna("").astype(str).str.strip() if doc_c else ""
                    out["_source"]     = filename
                    out["_sheet"]      = chosen_name
                    out = out[out["_rama"].str.len() > 2].reset_index(drop=True)
                    return out, None
                except Exception as e:
                    import traceback
                    return None, f"{e}\n{traceback.format_exc()}"

            # ── Parse uploaded files ──────────────────────────────────────────────────
            if xf_uploads:
                new_frames = []
                for uf in xf_uploads:
                    raw_bytes = uf.read()
                    parsed, err = _parse_facility(raw_bytes, uf.name)
                    if parsed is not None and len(parsed) > 0:
                        new_frames.append(parsed)
                        st.success(f"✅ **{uf.name}** — {len(parsed):,} visit records loaded "
                                   f"(sheet: *{parsed['_sheet'].iloc[0]}*)")
                    else:
                        st.error(f"❌ **{uf.name}** — {err}")
                if new_frames:
                    st.session_state["fd_facility"] = pd.concat(new_frames, ignore_index=True)

            if "fd_facility" not in st.session_state:
                st.info("👆 Upload at least one hospital or clinic file to run fraud detection.")
                st.stop()

            fac_df = st.session_state["fd_facility"]
            fac_ramas = set(fac_df["_rama"].tolist())

            # File summary
            source_summary = fac_df.groupby("_source").size().reset_index(name="records")
            cols_fsum = st.columns(min(len(source_summary), 4))
            for i, row in source_summary.iterrows():
                cols_fsum[i % len(cols_fsum)].metric(
                    row["_source"][:35], f"{row['records']:,} records"
                )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── STEP 2: Matching config ────────────────────────────────────────────────
            st.markdown('<div class="sec-head">⚙️ Step 2 — Detection Settings</div>',
                        unsafe_allow_html=True)

            cfg1, cfg2, cfg3 = st.columns(3)
            with cfg1:
                date_window = st.slider("Date window (±days)", 0, 30, 7,
                    help="How many days before/after a pharmacy dispensing date to look for a matching hospital visit")
            with cfg2:
                name_thresh = st.slider("Name similarity (0–1)", 0.0, 1.0, 0.4, 0.05,
                    help="Token overlap score between pharmacy and hospital patient names. Lower = more permissive.")
            with cfg3:
                require_name = st.checkbox("Require name match", value=True,
                    help="If OFF, matches on RAMA number alone (catches name typos)")

            # ── Build pharmacy working set ────────────────────────────────────────────
            def _ph_col(*keys):
                for k in keys:
                    if k in df.columns: return k
                return None

            vid_c = _ph_col("voucher_id","paper_code","Paper Code")
            pnm_c = _ph_col("patient_name","Patient Name")
            rma_c = _ph_col("patient_id","rama_number","RAMA Number")
            dt_c  = _ph_col("visit_date","dispensing_date","Dispensing Date")
            ins_c = _ph_col("insurance_copay","Insurance Co-payment")
            tot_c = _ph_col("amount","total_cost","Total Cost")
            doc_c = _ph_col("doctor_name","practitioner_name","Practitioner Name")
            dpt_c = _ph_col("practitioner_type","Practitioner Type")

            ph_work = df.copy()
            ph_work["_rama"]  = ph_work[rma_c].astype(str).str.strip().str.upper() if rma_c else ""
            ph_work["_name"]  = ph_work[pnm_c].fillna("").astype(str).str.strip()  if pnm_c else ""
            ph_work["_date"]  = pd.to_datetime(ph_work[dt_c], errors="coerce")     if dt_c  else pd.NaT
            ph_work["_vou"]   = ph_work[vid_c].astype(str).str.strip()             if vid_c else ""
            ph_work["_ins"]   = pd.to_numeric(ph_work[ins_c], errors="coerce").fillna(0) if ins_c else 0
            ph_work["_tot"]   = pd.to_numeric(ph_work[tot_c], errors="coerce").fillna(0) if tot_c else 0
            ph_work["_doc"]   = ph_work[doc_c].fillna("").astype(str)              if doc_c else ""
            ph_work["_dpt"]   = ph_work[dpt_c].fillna("").astype(str)              if dpt_c else ""

            def _tok(a, b):
                ta = set(str(a).upper().split()); tb = set(str(b).upper().split())
                return len(ta & tb) / len(ta | tb) if ta and tb else 0.0

            # ── Core matching ─────────────────────────────────────────────────────────
            # For each pharmacy row, find best facility match by RAMA + name + date
            results = []
            for _, pr in ph_work.iterrows():
                rama     = pr["_rama"]
                ph_date  = pr["_date"]
                ph_name  = pr["_name"]

                fac_rows = fac_df[fac_df["_rama"] == rama]

                if fac_rows.empty:
                    # NO facility record for this RAMA at all
                    results.append({
                        "status": "NO_RECORD",
                        "ph_voucher": pr["_vou"],
                        "ph_patient": ph_name,
                        "ph_rama":    rama,
                        "ph_date":    ph_date,
                        "ph_ins":     pr["_ins"],
                        "ph_total":   pr["_tot"],
                        "ph_doctor":  pr["_doc"],
                        "ph_dept":    pr["_dpt"],
                        "fac_voucher": None, "fac_name": None,
                        "fac_date":    None, "fac_source": None,
                        "days_apart":  None, "name_score": None,
                    })
                    continue

                # RAMA exists — check name + date
                best = None; best_delta = 9999; best_score = 0
                for _, fr in fac_rows.iterrows():
                    fac_date = fr["_date"]
                    nscore   = _tok(ph_name, fr["_name"])
                    delta    = abs((ph_date - fac_date).days) if pd.notna(ph_date) and pd.notna(fac_date) else 9999
                    name_ok  = (nscore >= name_thresh) if require_name else True
                    if name_ok and delta <= date_window:
                        if delta < best_delta or (delta == best_delta and nscore > best_score):
                            best_delta = delta; best_score = nscore; best = fr

                if best is not None:
                    # MATCHED — legitimate dispensing with traced visit
                    results.append({
                        "status":      "MATCHED",
                        "ph_voucher":  pr["_vou"],   "ph_patient": ph_name,
                        "ph_rama":     rama,          "ph_date":    ph_date,
                        "ph_ins":      pr["_ins"],    "ph_total":   pr["_tot"],
                        "ph_doctor":   pr["_doc"],    "ph_dept":    pr["_dpt"],
                        "fac_voucher": best["voucher_id"], "fac_name": best["_name"],
                        "fac_date":    best["_date"],      "fac_source": best["_source"],
                        "days_apart":  best_delta,    "name_score": round(best_score, 2),
                    })
                else:
                    # RAMA EXISTS but date/name mismatch — partial flag
                    fac_dates = fac_rows["_date"].dropna()
                    nearest_d = None
                    if not fac_dates.empty and pd.notna(ph_date):
                        deltas = (fac_dates - ph_date).abs()
                        nearest_d = int(deltas.min().days)
                    best_fr = fac_rows.iloc[0]
                    results.append({
                        "status":      "UNLINKED",
                        "ph_voucher":  pr["_vou"],   "ph_patient": ph_name,
                        "ph_rama":     rama,          "ph_date":    ph_date,
                        "ph_ins":      pr["_ins"],    "ph_total":   pr["_tot"],
                        "ph_doctor":   pr["_doc"],    "ph_dept":    pr["_dpt"],
                        "fac_voucher": best_fr["voucher_id"], "fac_name": best_fr["_name"],
                        "fac_date":    best_fr["_date"],      "fac_source": best_fr["_source"],
                        "days_apart":  nearest_d,     "name_score": round(_tok(ph_name, best_fr["_name"]),2),
                    })

            res_df = pd.DataFrame(results)

            no_rec    = res_df[res_df["status"]=="NO_RECORD"]
            unlinked  = res_df[res_df["status"]=="UNLINKED"]
            matched   = res_df[res_df["status"]=="MATCHED"]

            total_ins       = res_df["ph_ins"].sum()
            no_rec_ins      = no_rec["ph_ins"].sum()
            unlinked_ins    = unlinked["ph_ins"].sum()
            matched_ins     = matched["ph_ins"].sum()
            at_risk_ins     = no_rec_ins + unlinked_ins
            fac_count       = fac_df["_source"].nunique()
            coverage_pct    = 100 * matched_ins / total_ins if total_ins else 0

            # ── STEP 3: Dashboard ─────────────────────────────────────────────────────
            st.markdown('<div class="sec-head">📊 Step 3 — Fraud Detection Dashboard</div>',
                        unsafe_allow_html=True)

            # Top KPI strip
            k1,k2,k3,k4,k5 = st.columns(5)
            k1.metric("Total pharmacy vouchers", f"{len(ph_work):,}")
            k2.metric("✅ Verified (visit found)",
                      f"{len(matched):,}",
                      f"{100*len(matched)/len(ph_work):.1f}%")
            k3.metric("🔴 No facility record",
                      f"{len(no_rec):,}",
                      f"-{100*len(no_rec)/len(ph_work):.1f}%",
                      delta_color="inverse")
            k4.metric("🟡 RAMA found, visit unlinked",
                      f"{len(unlinked):,}",
                      f"-{100*len(unlinked)/len(ph_work):.1f}%",
                      delta_color="inverse")
            k5.metric("Facilities loaded", f"{fac_count}")

            st.markdown("<br>", unsafe_allow_html=True)

            # Insurance risk strip
            r1,r2,r3 = st.columns(3)
            r1.metric("Total RSSB claims (85%)",      f"RWF {total_ins:,.0f}")
            r2.metric("🔴 At-risk amount (no record)", f"RWF {no_rec_ins:,.0f}",
                      f"{100*no_rec_ins/total_ins:.1f}% of total",
                      delta_color="inverse")
            r3.metric("🟡 Partially unlinked",         f"RWF {unlinked_ins:,.0f}",
                      f"{100*unlinked_ins/total_ins:.1f}% of total",
                      delta_color="inverse")

            # Risk bar
            bar_matched  = 100 * matched_ins  / total_ins if total_ins else 0
            bar_unlinked = 100 * unlinked_ins / total_ins if total_ins else 0
            bar_norec    = 100 * no_rec_ins   / total_ins if total_ins else 0
            st.markdown(f"""
        <div style='margin:18px 0 8px'>
          <div style='font-size:11px;color:#64748b;font-family:monospace;margin-bottom:5px'>
            Insurance amount breakdown by verification status
          </div>
          <div style='display:flex;height:16px;border-radius:8px;overflow:hidden;width:100%'>
            <div style='width:{bar_matched:.1f}%;background:#22c55e' title='Verified: RWF {matched_ins:,.0f}'></div>
            <div style='width:{bar_unlinked:.1f}%;background:#f59e0b' title='Unlinked: RWF {unlinked_ins:,.0f}'></div>
            <div style='width:{bar_norec:.1f}%;background:#ef4444' title='No record: RWF {no_rec_ins:,.0f}'></div>
          </div>
          <div style='display:flex;gap:18px;margin-top:5px;font-size:11px;font-family:monospace'>
            <span style='color:#22c55e'>■ Verified {bar_matched:.1f}%</span>
            <span style='color:#f59e0b'>■ Unlinked {bar_unlinked:.1f}%</span>
            <span style='color:#ef4444'>■ No record {bar_norec:.1f}%</span>
          </div>
        </div>""", unsafe_allow_html=True)

            # Coverage note
            st.markdown(f"""
        <div style='background:#030d1a;border:1px solid #1e3a5f;border-radius:8px;
             padding:10px 16px;font-size:11px;font-family:monospace;color:#64748b;margin-bottom:20px'>
          <b style='color:#38bdf8'>Coverage note:</b>
          {fac_count} facility file(s) loaded covering {len(fac_ramas):,} unique RAMA numbers.
          Pharmacy serves patients from many facilities — patients in the
          <span style='color:#ef4444'>"No record"</span> group may have visited clinics
          <b>not yet uploaded</b>. Upload more facility files to reduce false positives.
        </div>""", unsafe_allow_html=True)

            st.markdown("<hr style='border-color:#1e2a38;margin:4px 0 24px'>", unsafe_allow_html=True)

            # ── TABLE 1: No Hospital Record ───────────────────────────────────────────
            st.markdown(f"""
        <div class='fraud-card fraud-card-red'>
          <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
            <div>
              <span style='font-size:16px;font-weight:800;color:#f87171;font-family:Syne,sans-serif'>
                🔴 Table 1 — No Hospital / Clinic Visit Record
              </span>
              <span class='badge badge-red' style='margin-left:10px'>{len(no_rec):,} vouchers</span>
              <span class='badge badge-red' style='margin-left:6px'>RWF {no_rec_ins:,.0f}</span>
            </div>
            <span style='font-size:11px;color:#991b1b;font-family:monospace'>
              Patient's RAMA number not found in ANY uploaded facility file
            </span>
          </div>
        </div>""", unsafe_allow_html=True)

            if not no_rec.empty:
                # Sub-controls
                t1c1, t1c2, t1c3 = st.columns([2,1.5,1.5])
                with t1c1:
                    t1_srch = st.text_input("🔍 Search", placeholder="Name, RAMA, voucher, doctor…",
                                             key="t1_srch")
                with t1c2:
                    t1_doc = st.selectbox("Filter by Prescriber",
                        ["All"] + sorted(no_rec["ph_doctor"].unique().tolist()),
                        key="t1_doc")
                with t1c3:
                    t1_min = st.number_input("Min insurance (RWF)", 0, value=0, step=5000, key="t1_min")

                t1_disp = no_rec.copy()
                if t1_srch:
                    mask = t1_disp.apply(
                        lambda c: c.astype(str).str.contains(t1_srch, case=False, na=False)
                    ).any(axis=1)
                    t1_disp = t1_disp[mask]
                if t1_doc != "All":
                    t1_disp = t1_disp[t1_disp["ph_doctor"] == t1_doc]
                if t1_min > 0:
                    t1_disp = t1_disp[t1_disp["ph_ins"] >= t1_min]

                # Build clean display table
                t1_show = t1_disp[[
                    "ph_voucher","ph_patient","ph_rama","ph_date",
                    "ph_ins","ph_total","ph_doctor","ph_dept"
                ]].copy()
                t1_show.columns = [
                    "Pharmacy Voucher","Patient Name","RAMA Number","Dispensing Date",
                    "Insurance Claim (RWF)","Total Cost (RWF)","Prescriber","Specialty"
                ]
                t1_show["Dispensing Date"] = pd.to_datetime(
                    t1_show["Dispensing Date"], errors="coerce"
                ).dt.strftime("%d/%m/%Y").fillna("—")
                t1_show = t1_show.sort_values("Insurance Claim (RWF)", ascending=False)
                t1_show.index = range(1, len(t1_show)+1)

                st.markdown(
                    f"<div style='font-size:11px;color:{MUTED};font-family:monospace;margin-bottom:6px'>"
                    f"Showing <b style='color:#f87171'>{len(t1_show):,}</b> vouchers · "
                    f"Insurance at risk: <b style='color:#ef4444'>RWF {t1_disp['ph_ins'].sum():,.0f}</b>"
                    f"</div>", unsafe_allow_html=True
                )
                st.dataframe(t1_show, use_container_width=True, height=340)

                # Prescriber risk breakdown
                with st.expander("📊 Prescriber risk breakdown (Table 1)", expanded=False):
                    doc_risk = (no_rec.groupby("ph_doctor")["ph_ins"]
                                .agg(Vouchers="count", Total_Claimed="sum")
                                .sort_values("Total_Claimed", ascending=False)
                                .reset_index())
                    doc_risk.columns = ["Prescriber","Vouchers","Total Claimed (RWF)"]
                    st.dataframe(doc_risk, use_container_width=True, height=280)

                # Download
                t1_buf = io.BytesIO()
                _t1_xl = no_rec.copy()
                _t1_xl["ph_date"] = pd.to_datetime(_t1_xl["ph_date"], errors="coerce").dt.strftime("%d/%m/%Y")
                with pd.ExcelWriter(t1_buf, engine="openpyxl") as xw:
                    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _Al
                    _t1_xl.to_excel(xw, index=False, sheet_name="No Facility Record")
                    ws = xw.sheets["No Facility Record"]
                    hf = _PF("solid", fgColor="7F1D1D")
                    for cell in ws[1]:
                        cell.fill = hf
                        cell.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                        cell.alignment = _Al(horizontal="center", wrap_text=True)
                    for i, r in enumerate(ws.iter_rows(min_row=2), 2):
                        bg = "FFE4E4" if i%2==0 else "FFFFFF"
                        for c in r: c.fill = _PF("solid", fgColor=bg)
                t1_buf.seek(0)
                st.download_button("⬇️ Download Table 1", t1_buf.getvalue(),
                    "table1_no_facility_record.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t1")

            st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

            # ── TABLE 2: UNLINKED (RAMA found, visit not linked) ──────────────────────
            st.markdown(f"""
        <div class='fraud-card fraud-card-amber'>
          <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
            <div>
              <span style='font-size:16px;font-weight:800;color:#fbbf24;font-family:Syne,sans-serif'>
                🟡 Table 2 — RAMA Found, Visit Not Linked
              </span>
              <span class='badge badge-amber' style='margin-left:10px'>{len(unlinked):,} vouchers</span>
              <span class='badge badge-amber' style='margin-left:6px'>RWF {unlinked_ins:,.0f}</span>
            </div>
            <span style='font-size:11px;color:#92400e;font-family:monospace'>
              Patient exists in a facility file but dispensing date is outside the ±{date_window}-day window
            </span>
          </div>
        </div>""", unsafe_allow_html=True)

            if not unlinked.empty:
                t2c1, t2c2 = st.columns([2,1])
                with t2c1:
                    t2_srch = st.text_input("🔍 Search", placeholder="Name, RAMA…", key="t2_srch")
                with t2c2:
                    t2_max_gap = st.number_input("Max days apart to show", 1, 365, 60, key="t2_gap")

                t2_disp = unlinked.copy()
                if t2_srch:
                    mask = t2_disp.apply(
                        lambda c: c.astype(str).str.contains(t2_srch, case=False, na=False)
                    ).any(axis=1)
                    t2_disp = t2_disp[mask]
                if t2_max_gap:
                    t2_disp = t2_disp[
                        t2_disp["days_apart"].isna() | (t2_disp["days_apart"] <= t2_max_gap)
                    ]

                t2_show = t2_disp[[
                    "ph_voucher","ph_patient","ph_rama","ph_date","ph_ins",
                    "fac_name","fac_date","fac_source","days_apart","name_score"
                ]].copy()
                t2_show.columns = [
                    "Pharmacy Voucher","Pharmacy Patient","RAMA","Pharmacy Date","Insurance (RWF)",
                    "Facility Patient","Facility Visit Date","Facility","Days Apart","Name Score"
                ]
                for dcol in ["Pharmacy Date","Facility Visit Date"]:
                    t2_show[dcol] = pd.to_datetime(t2_show[dcol], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")
                t2_show = t2_show.sort_values("Days Apart", na_position="last")
                t2_show.index = range(1, len(t2_show)+1)

                st.markdown(
                    f"<div style='font-size:11px;color:{MUTED};font-family:monospace;margin-bottom:6px'>"
                    f"Showing <b style='color:#fbbf24'>{len(t2_show):,}</b> vouchers · "
                    f"Insurance: <b style='color:#f59e0b'>RWF {t2_disp['ph_ins'].sum():,.0f}</b>"
                    f"</div>", unsafe_allow_html=True
                )
                st.dataframe(t2_show, use_container_width=True, height=300)

                t2_buf = io.BytesIO()
                _t2_xl = t2_disp.copy()
                _t2_xl["ph_date"]  = pd.to_datetime(_t2_xl["ph_date"],  errors="coerce").dt.strftime("%d/%m/%Y")
                _t2_xl["fac_date"] = pd.to_datetime(_t2_xl["fac_date"], errors="coerce").dt.strftime("%d/%m/%Y")
                with pd.ExcelWriter(t2_buf, engine="openpyxl") as xw:
                    from openpyxl.styles import PatternFill as _PF, Font as _F, Alignment as _Al
                    _t2_xl.to_excel(xw, index=False, sheet_name="Unlinked Visits")
                    ws = xw.sheets["Unlinked Visits"]
                    for cell in ws[1]:
                        cell.fill = _PF("solid", fgColor="78350F")
                        cell.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                        cell.alignment = _Al(horizontal="center", wrap_text=True)
                    for i, r in enumerate(ws.iter_rows(min_row=2), 2):
                        bg = "FEF3C7" if i%2==0 else "FFFFFF"
                        for c in r: c.fill = _PF("solid", fgColor=bg)
                t2_buf.seek(0)
                st.download_button("⬇️ Download Table 2", t2_buf.getvalue(),
                    "table2_unlinked_visits.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t2")

            st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

            # ── TABLE 3: MATCHED (verified, informational) ────────────────────────────
            st.markdown(f"""
        <div class='fraud-card fraud-card-green'>
          <div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px'>
            <div>
              <span style='font-size:16px;font-weight:800;color:#4ade80;font-family:Syne,sans-serif'>
                ✅ Table 3 — Verified: Hospital Visit + Pharmacy Dispensing Linked
              </span>
              <span class='badge badge-green' style='margin-left:10px'>{len(matched):,} vouchers</span>
              <span class='badge badge-green' style='margin-left:6px'>RWF {matched_ins:,.0f}</span>
            </div>
            <span style='font-size:11px;color:#14532d;font-family:monospace'>
              Legitimate patient journey confirmed — clinic visit → pharmacy dispensing
            </span>
          </div>
        </div>""", unsafe_allow_html=True)

            with st.expander("View verified records (Table 3)", expanded=False):
                t3_search = st.text_input("🔍 Search verified records", key="t3_srch")
                t3_show = matched[[
                    "ph_voucher","ph_patient","ph_rama","ph_date","ph_ins",
                    "fac_voucher","fac_name","fac_date","fac_source","days_apart","name_score"
                ]].copy()
                t3_show.columns = [
                    "Pharmacy Voucher","Pharmacy Patient","RAMA","Pharmacy Date","Insurance (RWF)",
                    "Facility Voucher","Facility Patient","Facility Date","Facility","Days Apart","Name Score"
                ]
                for dcol in ["Pharmacy Date","Facility Date"]:
                    t3_show[dcol] = pd.to_datetime(t3_show[dcol], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")
                if t3_search:
                    mask = t3_show.apply(
                        lambda c: c.astype(str).str.contains(t3_search, case=False, na=False)
                    ).any(axis=1)
                    t3_show = t3_show[mask]
                t3_show = t3_show.sort_values("Days Apart").reset_index(drop=True)
                t3_show.index = t3_show.index + 1
                st.dataframe(t3_show, use_container_width=True, height=300)

            st.markdown("<hr style='border-color:#1e2a38;margin:28px 0'>", unsafe_allow_html=True)

            # ── FULL REPORT DOWNLOAD ──────────────────────────────────────────────────
            st.markdown('<div class="sec-head">⬇️ Download Full Fraud Detection Report</div>',
                        unsafe_allow_html=True)

            if st.button("📊 Generate Full Report (4 sheets)", type="primary", key="fd_gen"):
                from openpyxl import Workbook as _WB
                from openpyxl.styles import (PatternFill as _PF, Font as _F,
                                             Alignment as _Al, Border as _B, Side as _S)
                from openpyxl.utils import get_column_letter as _gcl

                wb = _WB(); wb.remove(wb.active)
                THIN = _S(border_style="thin", color="CCCCCC")
                BDR  = _B(left=THIN,right=THIN,top=THIN,bottom=THIN)

                def _make_sheet(wb, title, data_df, hdr_color, row_colors):
                    ws = wb.create_sheet(title)
                    for ci, col in enumerate(data_df.columns, 1):
                        c = ws.cell(1, ci, col)
                        c.fill = _PF("solid", fgColor=hdr_color)
                        c.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
                        c.alignment = _Al(horizontal="center", wrap_text=True)
                        c.border = BDR
                        ws.column_dimensions[_gcl(ci)].width = max(14, min(len(str(col))+4, 35))
                    ws.row_dimensions[1].height = 30
                    ws.freeze_panes = "A2"
                    for ri, (_, row) in enumerate(data_df.iterrows(), 2):
                        bg = row_colors[ri % len(row_colors)]
                        for ci, val in enumerate(row, 1):
                            v = "" if (isinstance(val, float) and math.isnan(val)) else val
                            c = ws.cell(ri, ci, v)
                            c.font = _F(name="Arial", size=10)
                            c.fill = _PF("solid", fgColor=bg)
                            c.border = BDR
                            c.alignment = _Al(horizontal="left")
                    return ws

                # Sheet 0: Summary
                ws0 = wb.create_sheet("Summary")
                ws0.sheet_view.showGridLines = False
                summary_data = [
                    ("Pharmacy Report Period", "January 2025"),
                    ("Total Pharmacy Vouchers", len(ph_work)),
                    ("Facilities Loaded", fac_count),
                    ("", ""),
                    ("✅ Verified (visit found)", len(matched)),
                    ("   % of vouchers", f"{100*len(matched)/len(ph_work):.1f}%"),
                    ("   Insurance amount (RWF)", matched_ins),
                    ("", ""),
                    ("🔴 No Facility Record", len(no_rec)),
                    ("   % of vouchers", f"{100*len(no_rec)/len(ph_work):.1f}%"),
                    ("   Insurance at risk (RWF)", no_rec_ins),
                    ("", ""),
                    ("🟡 RAMA Found, Visit Unlinked", len(unlinked)),
                    ("   % of vouchers", f"{100*len(unlinked)/len(ph_work):.1f}%"),
                    ("   Insurance (RWF)", unlinked_ins),
                    ("", ""),
                    ("TOTAL INSURANCE AT RISK (RWF)", at_risk_ins),
                    ("As % of total claims", f"{100*at_risk_ins/total_ins:.1f}%"),
                ]
                for ri, (label, val) in enumerate(summary_data, 2):
                    lc = ws0.cell(ri, 1, label)
                    vc = ws0.cell(ri, 2, val)
                    lc.font = _F(name="Arial", size=11, bold=("TOTAL" in str(label) or "%" not in str(label) and str(label).startswith(("✅","🔴","🟡","Pharmacy","Facilities"))))
                    vc.font = _F(name="Arial", size=11, bold="TOTAL" in str(label))
                    if "TOTAL" in str(label):
                        lc.fill = _PF("solid", fgColor="7F1D1D")
                        vc.fill = _PF("solid", fgColor="7F1D1D")
                        lc.font = _F(name="Arial", size=12, bold=True, color="FFFFFF")
                        vc.font = _F(name="Arial", size=12, bold=True, color="FFFFFF")
                ws0.column_dimensions["A"].width = 38
                ws0.column_dimensions["B"].width = 22

                # Sheet 1: No record
                def _fmt_date(s):
                    return pd.to_datetime(s, errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
                t1_xl = no_rec[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","ph_total","ph_doctor","ph_dept"]].copy()
                t1_xl.columns = ["Voucher","Patient Name","RAMA Number","Dispensing Date","Insurance (RWF)","Total Cost (RWF)","Prescriber","Specialty"]
                t1_xl["Dispensing Date"] = _fmt_date(t1_xl["Dispensing Date"])
                _make_sheet(wb, "1 - No Facility Record", t1_xl, "7F1D1D", ["FFE4E4","FFFFFF"])

                # Sheet 2: Unlinked
                t2_xl = unlinked[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","fac_name","fac_date","fac_source","days_apart","name_score"]].copy()
                t2_xl.columns = ["Voucher","Patient Name","RAMA","Pharmacy Date","Insurance (RWF)","Facility Patient","Facility Date","Facility","Days Apart","Name Score"]
                t2_xl["Pharmacy Date"] = _fmt_date(t2_xl["Pharmacy Date"])
                t2_xl["Facility Date"] = _fmt_date(t2_xl["Facility Date"])
                _make_sheet(wb, "2 - Unlinked Visits", t2_xl, "78350F", ["FEF3C7","FFFFFF"])

                # Sheet 3: Verified
                t3_xl = matched[["ph_voucher","ph_patient","ph_rama","ph_date","ph_ins","fac_voucher","fac_name","fac_date","fac_source","days_apart","name_score"]].copy()
                t3_xl.columns = ["Voucher","Patient Name","RAMA","Pharmacy Date","Insurance (RWF)","Facility Voucher","Facility Patient","Facility Date","Facility","Days Apart","Name Score"]
                t3_xl["Pharmacy Date"] = _fmt_date(t3_xl["Pharmacy Date"])
                t3_xl["Facility Date"] = _fmt_date(t3_xl["Facility Date"])
                _make_sheet(wb, "3 - Verified", t3_xl, "14532D", ["E7F5EC","FFFFFF"])

                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                st.success("✅ Full report ready!")
                st.download_button(
                    "⬇️ Download Full Fraud Detection Report (.xlsx)",
                    buf.getvalue(), "fraud_detection_report.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_full_fd"
                )



