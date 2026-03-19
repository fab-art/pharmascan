"""
exporter.py — Excel report generation.

export_rules_excel():
    5-sheet formatted workbook from rules engine results.
    Sheet 1: Executive Summary
    Sheet 2: BLOCK (critical)
    Sheet 3: HOLD (review)
    Sheet 4: FLAG
    Sheet 5: Provider Risk Ranking

generate_counter_verification_xlsx():
    Official RSSB 2-sheet counter-verification report.
"""
import io
import math as _math
import pandas as pd

from utils import audit, LOG


def export_rules_excel(out_df: pd.DataFrame, summary: dict,
                       filename: str = "pharmascan_fraud_report.xlsx") -> bytes:
    """
    Generate a professional 5-sheet Excel report from rules engine output.
    Sheet 1: Executive Summary
    Sheet 2: BLOCK decisions (critical)
    Sheet 3: HOLD decisions
    Sheet 4: FLAG decisions
    Sheet 5: Provider Risk Ranking
    """
    from openpyxl import Workbook as _WB
    from openpyxl.styles import (
        Font as _F, PatternFill as _PF, Alignment as _Al,
        Border as _B, Side as _S,
    )
    from openpyxl.utils import get_column_letter as _gcl

    THIN = _S(border_style="thin", color="CCCCCC")
    BDR  = _B(left=THIN, right=THIN, top=THIN, bottom=THIN)

    PALETTE = {
        "BLOCK":   ("7F1D1D","FECACA"),
        "HOLD":    ("78350F","FEF3C7"),
        "FLAG":    ("1E3A5F","DBEAFE"),
        "APPROVE": ("14532D","DCFCE7"),
        "HEADER":  "003366",
        "GOLD":    "FFCC00",
    }

    wb = _WB()
    wb.remove(wb.active)

    def _hdr(ws, cols, fill_hex):
        for ci, (lbl, w) in enumerate(cols, 1):
            c = ws.cell(1, ci, lbl)
            c.fill = _PF("solid", fgColor=fill_hex)
            c.font = _F(bold=True, color="FFFFFF", name="Arial", size=10)
            c.alignment = _Al(horizontal="center", wrap_text=True)
            c.border = BDR
            ws.column_dimensions[_gcl(ci)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    def _data_row(ws, ri, vals, bg_hex):
        fill = _PF("solid", fgColor=bg_hex)
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci, val if not (isinstance(val, float) and _math.isnan(val)) else "")
            c.font = _F(name="Arial", size=10)
            c.fill = fill
            c.border = BDR
            c.alignment = _Al(horizontal="left", vertical="top", wrap_text=True)

    # ── Sheet 1: Executive Summary ─────────────────────────────────────
    ws0 = wb.create_sheet("Executive Summary")
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 38
    ws0.column_dimensions["B"].width = 20
    ws0.column_dimensions["C"].width = 16

    # Title
    ws0.merge_cells("A1:C2")
    t = ws0["A1"]
    t.value = "PharmaScan — Fraud Detection Report"
    t.font  = _F(bold=True, size=18, color="FFFFFF", name="Arial")
    t.fill  = _PF("solid", fgColor=PALETTE["HEADER"])
    t.alignment = _Al(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 28
    ws0.row_dimensions[2].height = 28

    d = summary.get("decisions", {})
    total = summary.get("total", 1) or 1
    rows_s0 = [
        ("","",""),
        ("Total claims evaluated", summary.get("total",0), ""),
        ("Rules engine version",   summary.get("rules_version","—"), ""),
        ("Processing time (ms)",   summary.get("elapsed_ms","—"), ""),
        ("","",""),
        ("✅ APPROVE",  d.get("APPROVE",0), f"{100*d.get('APPROVE',0)/total:.1f}%"),
        ("🟡 FLAG",     d.get("FLAG",0),    f"{100*d.get('FLAG',0)/total:.1f}%"),
        ("🟠 HOLD",     d.get("HOLD",0),    f"{100*d.get('HOLD',0)/total:.1f}%"),
        ("🔴 BLOCK",    d.get("BLOCK",0),   f"{100*d.get('BLOCK',0)/total:.1f}%"),
        ("","",""),
        ("Total flagged (FLAG+HOLD+BLOCK)", summary.get("flagged_count",0),
         f"{100*summary.get('flagged_count',0)/total:.1f}%"),
        ("At-risk insurance amount (RWF)",
         f"{summary.get('total_flagged_amount',0):,.2f}", ""),
    ]
    for ri, (lbl, val, pct) in enumerate(rows_s0, 3):
        ws0.cell(ri, 1, lbl).font  = _F(name="Arial", size=11,
                                         bold=any(x in lbl for x in ("TOTAL","APPROVE","FLAG","HOLD","BLOCK","Total","At-risk")))
        ws0.cell(ri, 2, val).font  = _F(name="Arial", size=11, bold=True)
        ws0.cell(ri, 3, pct).font  = _F(name="Arial", size=10, color="64748b")
        if "BLOCK" in lbl:
            for ci in (1,2,3):
                ws0.cell(ri, ci).fill = _PF("solid", fgColor="FECACA")
        elif "HOLD" in lbl:
            for ci in (1,2,3):
                ws0.cell(ri, ci).fill = _PF("solid", fgColor="FEF3C7")

    # Top rules fired
    ws0.cell(len(rows_s0)+4, 1, "Top Rules by Fires").font = _F(bold=True, name="Arial", size=11)
    for r_off, (rid, cnt) in enumerate(summary.get("rules_with_most_fires",[])[:8]):
        ws0.cell(len(rows_s0)+5+r_off, 1, rid)
        ws0.cell(len(rows_s0)+5+r_off, 2, cnt)

    # ── Sheets 2-4: BLOCK / HOLD / FLAG ───────────────────────────────
    cols_data = [
        ("Voucher ID",     16), ("Patient ID",    18), ("Drug Code",     16),
        ("Diagnosis",      12), ("Doctor Type",   18), ("Amount (RWF)",  16),
        ("Score",           8), ("Decision",      10), ("Rules Fired",   28),
        ("Reasons",        55),
    ]
    _vou_c  = next((c for c in ["voucher_id"] if c in out_df.columns), None)
    _pid_c  = next((c for c in ["patient_id","patient_name"] if c in out_df.columns), None)
    _drg_c  = next((c for c in ["drug_code"] if c in out_df.columns), None)
    _dx_c   = next((c for c in ["diagnosis"] if c in out_df.columns), None)
    _doc_c  = next((c for c in ["doctor_type","doctor_name"] if c in out_df.columns), None)
    _amt_c2 = next((c for c in ["insurance_copay","amount"] if c in out_df.columns), None)
    result_cols = [c for c in [
        _vou_c, _pid_c, _drg_c, _dx_c, _doc_c, _amt_c2,
        "_score","_decision","_rules_fired","_reasons",
    ] if c]

    for decision_key, sheet_title in [
        ("BLOCK","🔴 Critical (BLOCK)"),
        ("HOLD", "🟠 Review (HOLD)"),
        ("FLAG", "🟡 Flagged (FLAG)"),
    ]:
        subset = out_df[out_df["_decision"] == decision_key].copy()
        if subset.empty:
            continue
        subset = subset.sort_values("_score", ascending=False)
        ws = wb.create_sheet(sheet_title[:31])
        ws.sheet_view.showGridLines = False

        hdr_hex, row_hex = PALETTE[decision_key]
        _hdr(ws, cols_data, hdr_hex)
        for ri, (_, row) in enumerate(subset[
            [c for c in result_cols if c in subset.columns]
        ].iterrows(), 2):
            bg = row_hex if (ri % 2 == 0) else "FFFFFF"
            vals = [row.get(c,"") for c in result_cols if c in subset.columns]
            _data_row(ws, ri, vals, bg)

    # ── Sheet 5: Provider Risk Ranking ────────────────────────────────
    doc_col_e = next((c for c in ["doctor_type","doctor_name"] if c in out_df.columns), None)
    amt_col_e = next((c for c in ["insurance_copay","amount"]  if c in out_df.columns), None)
    if doc_col_e:
        ws5 = wb.create_sheet("Provider Risk Ranking")
        ws5.sheet_view.showGridLines = False
        prov = (
            out_df.groupby(doc_col_e).agg(
                total_claims   = ("_score","count"),
                avg_score      = ("_score","mean"),
                max_score      = ("_score","max"),
                flagged        = ("_decision", lambda x: (x.isin(["FLAG","HOLD","BLOCK"])).sum()),
                blocked        = ("_decision", lambda x: (x == "BLOCK").sum()),
                total_amount   = (amt_col_e, "sum") if amt_col_e else ("_score","count"),
            ).reset_index()
            .assign(flag_rate=lambda d: (100 * d["flagged"] / d["total_claims"].clip(lower=1)).round(1))
            .sort_values("flag_rate", ascending=False)
        )
        prov["avg_score"] = prov["avg_score"].round(1)
        prov["total_amount"] = prov["total_amount"].round(0)

        p_cols = [
            (doc_col_e.replace("_"," ").title(), 30),
            ("Total Claims", 14), ("Avg Score", 12), ("Max Score", 12),
            ("Flagged",      12), ("Blocked",   12),
            ("Flag Rate %",  14), ("Total Amount (RWF)", 20),
        ]
        _hdr(ws5, p_cols, PALETTE["HEADER"])
        for ri, (_, row) in enumerate(prov.iterrows(), 2):
            bg = "FEF3C7" if row["flag_rate"] >= 40 else \
                 "FEF9C3" if row["flag_rate"] >= 20 else "FFFFFF"
            _data_row(ws5, ri, list(row), bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    audit("EXPORT_EXCEL", f"5-sheet report, {len(out_df):,} rows", len(out_df))
    return buf.read()



def generate_counter_verification_xlsx(
    df: pd.DataFrame,
    deductions: list[dict],
    meta: dict,
    prepared_by: str,
    verified_by: str,
    approved_by: str,
    pc_col:  str | None = None,
    ins_col: str | None = None,
    tot_col: str | None = None,
    obs_col: str | None = None,
    dif_col: str | None = None,
) -> bytes:
    """
    Generate counter-verification report.
    All dimensions matched exactly from reference file.

    Sheet 1 — "After counter verification"
    Sheet 2 — "Counter verification report"
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── Palette (exact hex from reference file) ───────────────────────────────
    C_BLUE      = "003366"   # Primary Blue
    C_GOLD      = "FFCC00"   # Accent Gold
    C_GREY      = "F4F4F4"   # Zebra / metadata bg
    C_WHITE     = "FFFFFF"
    C_TEXT      = "333333"   # Body text
    C_RED       = "C0392B"   # Deduction amounts / NO amounts
    C_GREEN     = "1E7E34"   # Verified YES text
    C_AMBER     = "B8860B"   # Verified NO text
    C_FILL_GREEN= "D4EDDA"   # YES background
    C_FILL_AMBER= "FFF3CD"   # NO background
    C_TITLE_BG  = "E8F0F7"   # Title banner background

    # ── Style helpers ─────────────────────────────────────────────────────────
    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _font(name="Calibri", bold=False, size=11.0, color=C_TEXT):
        return Font(name=name, bold=bold, size=size, color=color)

    def side(style, color="000000"):
        return Side(border_style=style, color=color)

    THIN_GREY = side("thin",   "AAAAAA")
    MED_BLUE  = side("medium", C_BLUE)
    MED_GOLD  = side("medium", C_GOLD)
    THIN_ANY  = side("thin")          # thin with default colour (matches file's thin/?)
    NONE_S    = Side(border_style=None)

    def border_hdr_mid():
        """Header cell middle — medium blue all, gold bottom."""
        return Border(left=MED_BLUE, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_first():
        """Header cell column A/first — thin left, medium rest."""
        return Border(left=THIN_ANY, right=MED_BLUE, top=MED_BLUE, bottom=MED_GOLD)

    def border_hdr_last():
        """Header cell last column — medium left, thin right."""
        return Border(left=MED_BLUE, right=THIN_ANY, top=MED_BLUE, bottom=MED_GOLD)

    def border_data():
        """Regular data cell borders."""
        return Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_first():
        """Data cell col A — thin/? left."""
        return Border(left=THIN_ANY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    def border_data_last():
        """Data cell last col E — thin/? right."""
        return Border(left=THIN_GREY, right=THIN_ANY, top=THIN_GREY, bottom=THIN_GREY)

    # ── Alignments ────────────────────────────────────────────────────────────
    A_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_CENTER_NW = Alignment(horizontal="center", vertical="center", wrap_text=False)
    A_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    A_LEFT_NW   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    A_LEFT_TOP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    A_RIGHT     = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    # ── Numeric helpers ───────────────────────────────────────────────────────
    def _safe_float(v):
        if v is None: return 0.0
        try:
            if pd.isna(v): return 0.0
        except Exception: pass
        try: return float(str(v).replace(",", "").replace(" ", ""))
        except ValueError: return 0.0

    def _safe_date(v):
        if v is None: return None
        try:
            if pd.isna(v): return None
        except Exception: pass
        return v

    def _get_col(row_, *keys, default=""):
        for k in keys:
            if k and k in row_.index:
                v = row_[k]
                try:
                    if pd.notna(v): return v
                except Exception:
                    if v is not None: return v
        return default

    # ── Deduction lookup ──────────────────────────────────────────────────────
    ded_map = {str(d["paper_code"]).strip(): d for d in deductions}

    wb = Workbook()

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 — "After counter verification"
    # ═════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "After counter verification"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    # Column widths — exact from reference file
    col_widths_s1 = [
        13.54,  # A  Paper Code
        17.91,  # B  Dispensing Date
        22.54,  # C  Patient Name
        24.82,  # D  RAMA Number
        28.36,  # E  Practitioner Name
        18.18,  # F  Health Facility
        19.73,  # G  Date of Treatment
        13.63,  # H  Verified
        23.45,  # I  Total Before Counter-V
        35.82,  # J  85% After Counter-V
        23.36,  # K  After Counter-V 100%
        20.82,  # L  After Counter-V 85%
        19.73,  # M  Amount Deducted
        28.00,  # N  Explanation
    ]
    for ci, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 1: Header ─────────────────────────────────────────────────────────
    # font: Calibri 11 bold white; fill: #003366;
    # border: medium/003366 all sides, bottom medium/FFCC00
    headers_s1 = [
        "Paper Code", "Dispensing Date", "Patient Name", "RAMA Number",
        "Practitioner Name", "Health Facility", "Date of Treatment", "Verified",
        "Total Before Counter-V (RWF)", "85% After Counter-V (RWF)",
        "After Counter-V 100%", "After Counter-V 85%",
        "Amount Deducted (RWF)", "Explanation",
    ]
    for ci, h in enumerate(headers_s1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font      = _font("Calibri", bold=True, size=11.0, color=C_WHITE)
        c.fill      = fill(C_BLUE)
        c.border    = border_hdr_mid()
        c.alignment = A_CENTER
    ws1.row_dimensions[1].height = 33.75

    # ── Data validation: YES / NO dropdown col H ──────────────────────────────
    dv_verified = DataValidation(
        type="list", formula1='"YES,NO,PENDING"', allow_blank=True,
        showDropDown=False, showErrorMessage=True,
        errorTitle="Invalid value", error="Please choose YES, NO or PENDING",
    )
    ws1.add_data_validation(dv_verified)

    # ── Data rows ─────────────────────────────────────────────────────────────
    # font: Aptos Narrow 11 normal #333333; border: thin/AAAAAA all sides
    # row height: 46.25 (exact from reference)
    # zebra: even data_count (0-based) → white; odd → grey
    FONT_DATA  = _font("Aptos Narrow", bold=False, size=11.0, color=C_TEXT)
    FONT_YES   = _font("Aptos Narrow", bold=True,  size=11.0, color=C_GREEN)
    FONT_NO    = _font("Aptos Narrow", bold=True,  size=11.0, color=C_AMBER)
    FONT_RED   = _font("Aptos Narrow", bold=False, size=11.0, color=C_RED)

    data_count = 0
    for _, row in df.iterrows():
        ri = data_count + 2

        if pc_col and pc_col in row.index:
            pc = str(row[pc_col]).strip()
        else:
            pc = str(_get_col(row, "voucher_id", "paper_code", "Paper Code", default=""))

        ins_co     = _safe_float(_get_col(row, ins_col, "Insurance Co-payment",
                                          "insurance_copay", "insurance_co_payment"))
        ded        = ded_map.get(pc)
        is_ded     = ded is not None
        ded_amount = _safe_float(ded["amount"])      if is_ded else 0.0
        expla      = str(ded["explanation"]).strip()  if is_ded else ""
        verified   = "NO" if is_ded else "YES"

        # ded_amount is negative (e.g. -11342); use abs for arithmetic
        total_85  = round(ins_co * 0.85, 2)
        after_100 = round(ins_co - abs(ded_amount), 2)
        after_85  = round(after_100 * 0.85, 2)

        row_vals = [
            pc,
            _safe_date(_get_col(row, "Dispensing Date", "dispensing_date", "visit_date")),
            str(_get_col(row, "Patient Name",       "patient_name",      default="")),
            str(_get_col(row, "RAMA Number",        "rama_number", "patient_id", default="")),
            str(_get_col(row, "Practitioner Name",  "practitioner_name", "doctor_name", default="")),
            str(_get_col(row, "Health Facility",    "Health facility",   "facility",
                         default="PHARMACIE VINCA GISENYI LTD")),
            _safe_date(_get_col(row, "Dispensing Date", "dispensing_date", "visit_date")),
            verified,
            ins_co,
            total_85,
            after_100,
            after_85,
            ded_amount,
            expla,
        ]

        row_fill = fill(C_WHITE) if data_count % 2 == 0 else fill(C_GREY)
        ws1.row_dimensions[ri].height = 46.25

        for ci, val in enumerate(row_vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = border_data()

            if ci == 8:  # H — Verified
                c.font      = FONT_YES if verified == "YES" else FONT_NO
                c.fill      = fill(C_FILL_GREEN) if verified == "YES" else fill(C_FILL_AMBER)
                c.alignment = A_CENTER
                dv_verified.add(c)
            elif ci in (9, 10, 11, 12):  # Numeric amount cols
                c.font          = FONT_DATA
                c.fill          = row_fill
                c.number_format = "#,##0.00"
                c.alignment     = A_RIGHT
            elif ci == 13:  # M — Amount Deducted
                c.font          = FONT_RED if ded_amount != 0 else FONT_DATA
                c.fill          = row_fill
                c.number_format = "#,##0.00"
                c.alignment     = A_RIGHT
            elif ci in (2, 7):  # Date cols
                c.font          = FONT_DATA
                c.fill          = row_fill
                c.number_format = "dd/mm/yyyy"
                c.alignment     = A_CENTER
            elif ci == 14:  # N — Explanation
                c.font      = FONT_DATA
                c.fill      = row_fill
                c.alignment = A_LEFT
            else:
                c.font      = FONT_DATA
                c.fill      = row_fill
                c.alignment = A_LEFT

        data_count += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    # font: Calibri 11 bold white; fill: #003366
    # border: medium/003366 all sides (top = gold from reference, bottom = blue)
    if data_count > 0:
        tot_ri = data_count + 2
        ws1.row_dimensions[tot_ri].height = 21.75
        FONT_TOT = _font("Calibri", bold=True, size=11.0, color=C_WHITE)
        tot_border = Border(left=MED_BLUE, right=MED_BLUE, top=MED_GOLD, bottom=MED_BLUE)
        for ci in range(1, 15):
            c = ws1.cell(row=tot_ri, column=ci)
            c.font   = FONT_TOT
            c.fill   = fill(C_BLUE)
            c.border = tot_border
        ws1.cell(row=tot_ri, column=8, value="TOTAL").alignment = A_CENTER
        for ci, col_letter in [(9, "I"), (10, "J"), (11, "K"), (12, "L"), (13, "M")]:
            c = ws1.cell(row=tot_ri, column=ci,
                         value=f"=SUM({col_letter}2:{col_letter}{tot_ri - 1})")
            c.number_format = "#,##0.00"
            c.alignment     = A_RIGHT

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 — "Counter verification report"
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Counter verification report")
    ws2.sheet_view.showGridLines = False

    # Column widths — exact from reference file
    ws2.column_dimensions["A"].width = 12.24
    ws2.column_dimensions["B"].width = 25.04
    ws2.column_dimensions["C"].width = 31.86
    ws2.column_dimensions["D"].width = 13.49
    ws2.column_dimensions["E"].width = 49.24

    # ── Rows 1-3: Title banner — A1:E3 merged ────────────────────────────────
    # font: Calibri 36 bold #003366; fill: #E8F0F7; h=center v=center wrap=True
    ws2.merge_cells("A1:E3")
    t = ws2["A1"]
    t.value     = "RSSB - COUNTER VERIFICATION REPORT"
    t.font      = _font("Calibri", bold=True, size=36.0, color=C_BLUE)
    t.fill      = fill(C_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    t.border    = Border(left=THIN_ANY, right=THIN_ANY, top=THIN_ANY, bottom=NONE_S)
    for rn in [1, 2, 3]:
        ws2.row_dimensions[rn].height = 21.75

    # ── Row 4: Gold separator — height 3.75 ───────────────────────────────────
    ws2.row_dimensions[4].height = 3.75
    for ci in range(1, 6):
        ws2.cell(row=4, column=ci).fill = fill(C_GOLD)

    # ── Rows 5-9: Metadata ────────────────────────────────────────────────────
    # Label: col A — Calibri 13 bold #003366, h=left v=center, bL=thin
    # Value: col C — Calibri 13 bold #333333, h=left v=center
    # Col E last: bR=thin (right-side border of the block)
    meta_rows = [
        (5, "PROVINCE:",               meta.get("province", "")),
        (6, "ADMINISTRATIVE DISTRICT:", meta.get("district", "")),
        (7, "PHARMACY:",               meta.get("pharmacy", "")),
        (8, "PERIOD:",                 meta.get("period",   "")),
        (9, "CODE:",                   meta.get("code",     "")),
    ]
    for rn, label, value in meta_rows:
        ws2.row_dimensions[rn].height = 21.75
        lc = ws2.cell(row=rn, column=1, value=label)
        lc.font      = _font("Calibri", bold=True, size=13.0, color=C_BLUE)
        lc.fill      = fill(C_WHITE)
        lc.border    = Border(left=THIN_ANY)
        lc.alignment = A_LEFT_NW

        vc = ws2.cell(row=rn, column=3, value=value)
        vc.font      = _font("Calibri", bold=True, size=13.0, color=C_TEXT)
        vc.fill      = fill(C_WHITE)
        vc.alignment = A_LEFT_NW

        # Right-side border on col E
        ws2.cell(row=rn, column=5).border = Border(right=THIN_ANY)
        ws2.cell(row=rn, column=5).fill   = fill(C_WHITE)

    # ── Row 10: Spacer ────────────────────────────────────────────────────────
    ws2.row_dimensions[10].height = 9.75
    ws2.cell(row=10, column=1).border = Border(left=THIN_ANY)
    ws2.cell(row=10, column=5).border = Border(right=THIN_ANY)

    # ── Row 11: Table header ──────────────────────────────────────────────────
    # font: Calibri 10.5 bold white; fill: #003366; height: 36
    # A:  bL=thin, bR=medium/blue, bT=medium/blue, bBot=medium/gold
    # B-D: bL=medium/blue, bR=medium/blue, bT=medium/blue, bBot=medium/gold
    # E:  bL=medium/blue, bR=thin, bT=medium/blue, bBot=medium/gold
    ws2.row_dimensions[11].height = 36.0
    tbl_headers = [
        "No.", "Invoice ID\n(Paper Code)",
        "Beneficiary RAMA No.",
        "Amount Deducted (RWF)",
        "Explanation of Deduction",
    ]
    hdr_borders = [
        border_hdr_first(),  # A
        border_hdr_mid(),    # B
        border_hdr_mid(),    # C
        border_hdr_mid(),    # D
        border_hdr_last(),   # E
    ]
    for ci, (h, bdr) in enumerate(zip(tbl_headers, hdr_borders), 1):
        c = ws2.cell(row=11, column=ci, value=h)
        c.font      = _font("Calibri", bold=True, size=10.5, color=C_WHITE)
        c.fill      = fill(C_BLUE)
        c.border    = bdr
        c.alignment = A_CENTER

    # ── Rows 12+: Deduction data rows ─────────────────────────────────────────
    # font: Aptos Narrow 10.5 normal #333333; height: 18.0
    # zebra: i=0,2,4 → white; i=1,3,5 → grey
    # A: h=center v=center wrap, bL=thin/?, bR=thin/AAA, bT=thin/AAA, bBot=thin/AAA
    # B,C: h=left v=center wrap, thin/AAA all
    # D: h=right v=center no-wrap, color=red C0392B, numfmt=#,##0, thin/AAA all
    # E: h=left v=top wrap, bL=thin/AAA, bR=thin/?, bT=thin/AAA, bBot=thin/AAA
    n_ded      = len(deductions)
    n_rows     = n_ded          # rows grow with data — no padding
    data_start = 12

    FONT_D2 = _font("Aptos Narrow", bold=False, size=10.5, color=C_TEXT)
    FONT_DR = _font("Aptos Narrow", bold=False, size=10.5, color=C_RED)

    for i, d in enumerate(deductions):
        ri = data_start + i
        ws2.row_dimensions[ri].height = 18.0
        row_fill = fill(C_WHITE) if i % 2 == 0 else fill(C_GREY)

        vals = [
            i + 1,
            str(d.get("paper_code",  "")),
            str(d.get("rama_no",     "")),
            _safe_float(d.get("amount", 0)),
            str(d.get("explanation", "")),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.fill = row_fill

            if ci == 1:
                c.font      = FONT_D2
                c.border    = border_data_first()
                c.alignment = A_CENTER
            elif ci in (2, 3):
                c.font      = FONT_D2
                c.border    = border_data()
                c.alignment = A_LEFT
            elif ci == 4:
                c.font          = FONT_DR
                c.border        = border_data()
                c.number_format = "#,##0;-#,##0"
                c.alignment     = A_RIGHT
            elif ci == 5:
                c.font      = FONT_D2
                c.border    = border_data_last()
                c.alignment = A_LEFT_TOP

    # ── Total row ─────────────────────────────────────────────────────────────
    # font: Calibri 10.5 bold white; fill: #003366; height: 21.75
    # A: bL=thin, bBot=medium/blue
    # B: bL=medium/blue, bBot=medium/blue
    # C: h=right v=center, value="TOTAL AMOUNT DEDUCTED", bL=medium/blue, bBot=medium/blue
    # D: h=right v=center, numfmt=#,##0, formula, bL=medium/blue, bBot=medium/blue
    # E: bL=medium/blue, bBot=medium/blue
    tot_row = data_start + n_rows
    ws2.row_dimensions[tot_row].height = 21.75
    FONT_TOT2 = _font("Calibri", bold=True, size=10.5, color=C_WHITE)
    tot_borders = [
        Border(left=THIN_ANY,  bottom=MED_BLUE),  # A
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # B
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # C
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # D
        Border(left=MED_BLUE,  bottom=MED_BLUE),  # E
    ]
    for ci, bdr in enumerate(tot_borders, 1):
        c = ws2.cell(row=tot_row, column=ci)
        c.font   = FONT_TOT2
        c.fill   = fill(C_BLUE)
        c.border = bdr

    ws2.cell(row=tot_row, column=3,
             value="TOTAL AMOUNT DEDUCTED").alignment = A_RIGHT
    tot_c = ws2.cell(row=tot_row, column=4,
                     value=f"=SUM(D{data_start}:D{tot_row - 1})")
    tot_c.number_format = "#,##0;-#,##0"
    tot_c.alignment     = A_RIGHT

    # ── Signature block ───────────────────────────────────────────────────────
    # Spacer row after total: height=12
    # Rows +1 to +5: cols A, C, E; height=19.5
    # Rows +6 to +10: height=15 (empty trailing rows)
    sig_start = tot_row + 2
    ws2.row_dimensions[sig_start - 1].height = 12.0

    for rn in range(sig_start, sig_start + 5):
        ws2.row_dimensions[rn].height = 19.5
    for rn in range(sig_start + 5, sig_start + 10):
        ws2.row_dimensions[rn].height = 15.0

    # Row 0 (+0): "Prepared by" / "Verified by" / "Approved By" — bold blue, underline
    sig_data = [
        # (offset, [(col, value, bold, color)])
        (0, [(1, "Prepared by",  True,  C_BLUE),
             (3, "Verified by",  True,  C_BLUE),
             (5, "Approved By",  True,  C_BLUE)]),
        (1, [(1, "Date:",        False, C_TEXT),
             (3, "Date:",        False, C_TEXT),
             (5, "Date:",        False, C_TEXT)]),
        (2, [(1, "Signature:",   False, C_TEXT),
             (3, "Signature:",   False, C_TEXT),
             (5, "Signature:",   False, C_TEXT)]),
        (3, [(1, "Names:",       False, C_TEXT),
             (3, f"Names:  {verified_by}", True, C_BLUE),
             (5, "Names:",       False, C_TEXT)]),
        (4, [(3, "Lead, Counter Verification", False, C_TEXT),
             (5, "Responsible of Pharmacy",    False, C_TEXT)]),
    ]
    for offset, cells in sig_data:
        rn = sig_start + offset
        for col, text, bold, color in cells:
            c = ws2.cell(row=rn, column=col, value=text)
            c.font      = _font("Calibri", bold=bold, size=10.5, color=color)
            c.alignment = A_LEFT_NW
            if offset == 0:
                c.border = Border(bottom=THIN_ANY)

    # ── Output ────────────────────────────────────────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()




